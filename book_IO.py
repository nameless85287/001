from openpyxl import load_workbook
import pandas as pd

#建立一個輸入數值(應為ISBN)偵測是否為10碼13碼並傳出的函式
def input_ISBN():
    book_ISBN=input("請輸入書籍ISBN(10碼或13碼)：")
    if len(book_ISBN) != 10 and len(book_ISBN) != 13:
        for item in range(3):
            print("不為10碼或13碼")
            book_ISBN = input("請重新輸入ISBN:")
            if len(book_ISBN) == 10 or len(book_ISBN) == 13:
                return book_ISBN
            else:
                print("請下次再來！")
                return None
    else:
        return book_ISBN

#印出每個輸入值
def print_book_info(book_name,book_writer,book_publisher,
            book_publish_year_month,book_ISBN):
    print(f"書名：{book_name}\n"+
          f"作者：{book_writer}\n"+
          f"出版社：{book_publisher}\n"+
          f"出版年月：{book_publish_year_month}\n"+
          f"ISBN:{book_ISBN}")

#將資料寫入EXCEL
def write_data_to_excel(book_name,book_writer,book_publisher,
            book_publish_year_month,book_ISBN):
    df=pd.read_excel("藏書清冊.xlsx")
    wb=load_workbook("藏書清冊.xlsx")
    ws=wb["Sheet1"]

    row=len(df)+2
    str_row=str(row)
    ws["A"+str_row]=book_name
    ws["B"+str_row]=book_writer
    ws["C"+str_row]=book_publisher
    ws["D"+str_row]=book_publish_year_month
    ws["E"+str_row]=book_ISBN
    wb.save("藏書清冊.xlsx")
